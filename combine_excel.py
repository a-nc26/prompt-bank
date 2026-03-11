#!/usr/bin/env python3
"""Combine multiple Excel files into one sheet with Source and Sheet columns."""

import openpyxl
from pathlib import Path

INPUT_FILES = [
    "/Users/avilurie/Downloads/AI Companions - September 2025.xlsx",
    "/Users/avilurie/Downloads/Gemini - Red Team Exercise.xlsx",
    "/Users/avilurie/Downloads/Meta AI Red Team - May 2025.xlsx",
    "/Users/avilurie/Downloads/Snap - My AI Safety Assessment.xlsx",
    "/Users/avilurie/Downloads/Enterprise Test 2025 (1).xlsx",
    "/Users/avilurie/Downloads/Enterprise Test 2025.xlsx",
    "/Users/avilurie/Downloads/IBM - Oct 2025.xlsx",
    "/Users/avilurie/Downloads/Aruaco - Oct 2025.xlsx",
]

OUTPUT_FILE = "/Users/avilurie/Desktop/Prompt Bank/Combined_Sheets.xlsx"


def combine_excel_files():
    # First pass: find max columns and collect all rows with (source, sheet, row_values)
    all_rows = []
    max_cols = 0

    for path in INPUT_FILES:
        p = Path(path)
        if not p.exists():
            print(f"Skip (not found): {path}")
            continue
        name = p.name
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"Error opening {path}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                row_list = [v for v in row]
                max_cols = max(max_cols, len(row_list))
                all_rows.append((name, sheet_name, row_list))
        wb.close()

    if not all_rows:
        print("No data found.")
        return

    # Build header: Source, Sheet, Column 1, Column 2, ...
    header = ["Source", "Sheet"] + [f"Column {i+1}" for i in range(max_cols)]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Combined"

    for c, val in enumerate(header, start=1):
        ws_out.cell(row=1, column=c, value=val)

    for r, (source, sheet, values) in enumerate(all_rows, start=2):
        ws_out.cell(row=r, column=1, value=source)
        ws_out.cell(row=r, column=2, value=sheet)
        for c, val in enumerate(values, start=3):
            ws_out.cell(row=r, column=c, value=val)

    wb_out.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    print(f"Total rows (incl. header): {len(all_rows) + 1}")
    print(f"Columns: {len(header)}")


if __name__ == "__main__":
    combine_excel_files()
