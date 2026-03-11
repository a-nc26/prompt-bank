#!/usr/bin/env python3
"""
Extract input prompts + classification from the combined sources,
deduplicate, normalize labels, and optionally use AI to reduce slop.
Output: clean sheet ready for Prompt Bank (text, category, abuseArea, rating, source).
"""

import openpyxl
from pathlib import Path
from collections import defaultdict
import csv
import os
import uuid

INPUT_FILES = [
    "/Users/avilurie/Downloads/AI Companions - September 2025.xlsx",
    "/Users/avilurie/Downloads/Gemini - Red Team Exercise.xlsx",
    "/Users/avilurie/Downloads/Meta AI Red Team - May 2025.xlsx",
    "/Users/avilurie/Downloads/Snap - My AI Safety Assessment.xlsx",
    "/Users/avilurie/Downloads/Enterprise Test 2025 (1).xlsx",
    "/Users/avilurie/Downloads/Enterprise Test 2025.xlsx",
    "/Users/avilurie/Downloads/IBM - Oct 2025.xlsx",
    "/Users/avilurie/Downloads/Aruaco - Oct 2025.xlsx",
]

OUTPUT_FILE = "/Users/avilurie/Desktop/Prompt Bank/Curated_Prompts.xlsx"
OUTPUT_CSV = "/Users/avilurie/Desktop/Prompt Bank/Curated_Prompts.csv"

# Per-file schema: which header names mean prompt, classification, harm area, and multi-turn.
SCHEMA = {
    "prompt_headers": ["INPUT PROMPT", "Prompt / output", "Prompt / Output"],
    "classification_headers": ["VIOLATIVE", "Label", "Safe?", "Seesion Label"],
    "harm_area_headers": ["Harm area", "VERTICAL", "Theme"],
    "session_id_headers": ["Session ID"],
    "user_model_headers": ["User/Model", "User / Model"],
    "sequence_id_headers": ["Sequence Id", "Sequence ID"],
}


def find_col_indices(headers):
    """Return (prompt_col, classification_col, harm_area_col, session_id_col, user_model_col, sequence_id_col)."""
    headers_clean = [(str(h).strip() if h is not None else "").lower() for h in headers]

    def _find(names, require_min_len=2):
        for name in names:
            n = name.strip().lower()
            for j, h in enumerate(headers_clean):
                if len(h) < require_min_len:
                    continue
                if n in h or h in n:
                    return j
        return None

    prompt_col = _find(SCHEMA["prompt_headers"])
    classification_col = _find(SCHEMA["classification_headers"])
    harm_area_col = _find(SCHEMA["harm_area_headers"])
    # Require at least 5 chars so we don't match None/empty or "Comments" as "session id"
    session_id_col = _find(SCHEMA["session_id_headers"], require_min_len=5)
    user_model_col = _find(SCHEMA["user_model_headers"], require_min_len=5)
    sequence_id_col = _find(SCHEMA["sequence_id_headers"], require_min_len=5)
    return prompt_col, classification_col, harm_area_col, session_id_col, user_model_col, sequence_id_col


def normalize_rating(raw):
    """Map various classification labels to standard: violative, non-violative, safe, unsafe, unknown."""
    if raw is None or (isinstance(raw, str) and not str(raw).strip()):
        return "unknown"
    s = str(raw).strip().lower()
    if s in ("yes", "y", "1", "true", "violative", "violation", "unsafe", "positive", "hit"):
        return "violative"
    if s in ("no", "n", "0", "false", "non-violative", "safe", "negative", "clean"):
        return "non-violative"
    if "violat" in s or "unsafe" in s or "harm" in s:
        return "violative"
    if "safe" in s or "clean" in s or "non" in s:
        return "non-violative"
    return "unknown"


def normalize_prompt(text):
    """Strip and collapse whitespace for dedup key; return None if empty."""
    if text is None:
        return None
    t = str(text).strip()
    if not t:
        return None
    return " ".join(t.split())


def _role_from_user_model(val):
    """Return 'user' or 'assistant' from User/Model column."""
    if val is None:
        return "user"
    s = str(val).strip().lower()
    if "model" in s or "assistant" in s:
        return "assistant"
    return "user"


def extract_prompts():
    """Yield (prompt_text, classification, harm_area, source_name, turn_type, conversation).
    turn_type is 'single-turn' or 'multi-turn'; conversation is list of {role, content} or None.
    """
    for path in INPUT_FILES:
        p = Path(path)
        if not p.exists():
            continue
        name = p.name
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"Error opening {path}: {e}")
            continue
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [h for h in rows[0]]
            prompt_col, classification_col, harm_area_col, session_id_col, user_model_col, sequence_id_col = find_col_indices(headers)
            if prompt_col is None or classification_col is None:
                continue

            # Multi-turn: group by session and emit one record per conversation
            if session_id_col is not None and user_model_col is not None:
                sessions_by_id = defaultdict(list)
                current_sid = None
                for row in rows[1:]:
                    row_list = list(row) if row else []
                    if len(row_list) <= max(prompt_col, user_model_col, classification_col):
                        continue
                    sid = row_list[session_id_col] if session_id_col < len(row_list) else None
                    if sid is not None:
                        current_sid = sid
                    if current_sid is None:
                        continue
                    content = (row_list[prompt_col] or "")
                    if not str(content).strip():
                        continue
                    role = _role_from_user_model(row_list[user_model_col] if user_model_col < len(row_list) else None)
                    label = row_list[classification_col] if classification_col < len(row_list) else None
                    harm_val = row_list[harm_area_col] if harm_area_col is not None and harm_area_col < len(row_list) else None
                    harm_str = str(harm_val).strip() if harm_val else ""
                    sessions_by_id[current_sid].append({
                        "role": role, "content": str(content).strip(),
                        "label": label, "harm": harm_str,
                        "seq": row_list[sequence_id_col] if sequence_id_col is not None and sequence_id_col < len(row_list) else None,
                    })
                for sid, turns in sessions_by_id.items():
                    if not turns:
                        continue
                    # Sort by sequence if available, else keep row order
                    if turns[0].get("seq") is not None:
                        try:
                            turns.sort(key=lambda t: (t["seq"] or ""))
                        except Exception:
                            pass
                    conv = [{"role": t["role"], "content": t["content"]} for t in turns]
                    first_user = next((t["content"] for t in turns if t["role"] == "user"), turns[0]["content"])
                    # Use last row's label for whole conversation; prefer violative if any turn was
                    labels = [t["label"] for t in turns if t.get("label")]
                    rating = normalize_rating(labels[-1] if labels else None)
                    for lbl in labels:
                        if normalize_rating(lbl) == "violative":
                            rating = "violative"
                            break
                    harm_str = "; ".join(dict.fromkeys(t["harm"] for t in turns if t.get("harm")))
                    yield (normalize_prompt(first_user) or first_user[:200], rating, harm_str, name, "multi-turn", conv)
                continue

            # Single-turn: one record per row (use first non-empty cell as fallback if prompt column empty)
            for row in rows[1:]:
                row_list = list(row) if row else []
                if len(row_list) <= max(prompt_col, classification_col):
                    continue
                prompt_val = row_list[prompt_col]
                class_val = row_list[classification_col]
                harm_val = row_list[harm_area_col] if harm_area_col is not None and len(row_list) > harm_area_col else None
                prompt_norm = normalize_prompt(prompt_val)
                if not prompt_norm:
                    # Fallback: use first non-empty cell in row (many source rows have prompt elsewhere)
                    for cell in row_list:
                        if cell is not None and str(cell).strip():
                            prompt_norm = normalize_prompt(cell)
                            if prompt_norm and len(prompt_norm) > 10:
                                break
                if not prompt_norm:
                    continue
                rating = normalize_rating(class_val)
                harm_str = str(harm_val).strip() if harm_val is not None else ""
                yield (prompt_norm, rating, harm_str, name, "single-turn", None)
        wb.close()


def dedupe_and_curate(use_ai=False, no_dedupe=False):
    """Gather all prompts, dedupe single-turn by text (unless no_dedupe) and multi-turn by conversation hash."""
    import json
    # Single-turn: key = normalized text -> list of (rating, harm, source)  [or keep each row if no_dedupe]
    by_prompt = defaultdict(list)
    # Multi-turn: key = hash of conversation -> (first_user, rating, harm, source, conv)
    by_convo = {}
    for prompt_text, rating, harm_area, source, turn_type, conversation in extract_prompts():
        if turn_type == "multi-turn" and conversation:
            key = json.dumps(conversation, sort_keys=True)
            if key not in by_convo:
                by_convo[key] = (prompt_text, rating, harm_area, source, conversation)
            else:
                old = by_convo[key]
                new_rating = rating if rating == "violative" else old[1]
                new_harm = (old[2] + "; " + harm_area).strip("; ") if harm_area else old[2]
                by_convo[key] = (old[0], new_rating, new_harm, old[3] + "; " + source, old[4])
        else:
            if no_dedupe:
                by_prompt[prompt_text + "\n__id__" + str(len(by_prompt))] = [(rating, harm_area, source)]
            else:
                by_prompt[prompt_text].append((rating, harm_area, source))

    rows = []
    for prompt_text, entries in by_prompt.items():
        if no_dedupe:
            prompt_text = prompt_text.split("\n__id__")[0]
        ratings = [e[0] for e in entries]
        harm_areas = [e[1] for e in entries if e[1]]
        sources = list(dict.fromkeys(e[2] for e in entries))
        if "violative" in ratings:
            rating = "violative"
        elif "non-violative" in ratings:
            rating = "non-violative"
        else:
            rating = ratings[0] if ratings else "unknown"
        abuse_area = "; ".join(dict.fromkeys(harm_areas)) if harm_areas else ""
        source_str = "; ".join(sources)
        rows.append({
            "text": prompt_text,
            "category": "Trust & Safety",
            "abuseArea": abuse_area,
            "rating": rating,
            "source": source_str,
            "turnType": "single-turn",
            "conversation": None,
        })

    for _key, (prompt_text, rating, harm_area, source, conv) in by_convo.items():
        rows.append({
            "text": prompt_text,
            "category": "Trust & Safety",
            "abuseArea": harm_area,
            "rating": rating,
            "source": source,
            "turnType": "multi-turn",
            "conversation": conv,
        })

    if use_ai and os.environ.get("OPENAI_API_KEY"):
        rows = _ai_normalize_categories(rows)
    return rows


def _serialize_conversation(conv):
    """JSON string for conversation (for CSV/Excel)."""
    if not conv:
        return ""
    import json
    return json.dumps(conv, ensure_ascii=False)


def _ai_normalize_categories(rows):
    """Optional: call OpenAI to normalize abuseArea into a fixed taxonomy and drop near-duplicates."""
    try:
        import openai
        client = openai.OpenAI()
            # Batch in chunks to avoid rate limits
        out = []
        for i in range(0, len(rows), 50):
            batch = rows[i : i + 50]
            prompts_for_api = [r["text"][:500] for r in batch]
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{
                    "role": "user",
                    "content": "For each of these red-team/safety test prompts, reply with exactly one short harm category from: child_safety, hate_speech, violence, self_harm, sexual_content, harassment, misinformation, jailbreak, other. One category per line, same number of lines as input."
                    + "\n\n" + "\n".join(f"{j+1}. {p}" for j, p in enumerate(prompts_for_api)),
                }],
                temperature=0,
            )
            text = (resp.choices[0].message.content or "").strip()
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            for j, r in enumerate(batch):
                if j < len(lines):
                    r["abuseArea"] = lines[j].replace(".", "").strip()
                out.append(r)
        return out
    except Exception as e:
        print("AI normalization skipped:", e)
        return rows


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--ai", action="store_true", help="Use OpenAI to normalize abuse categories (set OPENAI_API_KEY)")
    ap.add_argument("--no-dedupe", action="store_true", help="Keep every row (no dedup); single-turn rows not merged by text")
    ap.add_argument("--stats", action="store_true", help="Print extraction counts per file/source")
    args = ap.parse_args()

    if args.stats:
        single, multi = 0, 0
        from collections import Counter
        by_source = Counter()
        for pt, rating, harm, source, turn_type, conv in extract_prompts():
            by_source[source] += 1
            if turn_type == "multi-turn":
                multi += 1
            else:
                single += 1
        print("Extracted: single-turn rows", single, "| multi-turn conversations", multi, "| total", single + multi)
        print("By source:", dict(by_source))

    rows = dedupe_and_curate(use_ai=args.ai, no_dedupe=args.no_dedupe)
    single_out = sum(1 for r in rows if r.get("turnType") == "single-turn")
    multi_out = sum(1 for r in rows if r.get("turnType") == "multi-turn")
    print(f"Curated: {len(rows)} total ({single_out} single-turn, {multi_out} multi-turn)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Curated"
    headers = ["id", "text", "category", "abuseArea", "rating", "turnType", "conversation", "source"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=str(uuid.uuid4())[:8])
        ws.cell(row=r, column=2, value=row["text"])
        ws.cell(row=r, column=3, value=row["category"])
        ws.cell(row=r, column=4, value=row["abuseArea"])
        ws.cell(row=r, column=5, value=row["rating"])
        ws.cell(row=r, column=6, value=row["turnType"])
        ws.cell(row=r, column=7, value=_serialize_conversation(row.get("conversation")))
        ws.cell(row=r, column=8, value=row["source"])

    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")

    # CSV for Prompt Bank app (rating: violative→unsafe, else→safe; include turnType and conversation)
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["text", "category", "abuseArea", "rating", "turnType", "conversation", "source"])
        for row in rows:
            app_rating = "unsafe" if row["rating"] == "violative" else "safe"
            writer.writerow([
                row["text"],
                row["category"],
                row["abuseArea"],
                app_rating,
                row["turnType"],
                _serialize_conversation(row.get("conversation")),
                row["source"],
            ])
    print(f"Saved: {OUTPUT_CSV} (use Upload CSV in the app to import)")


if __name__ == "__main__":
    main()
